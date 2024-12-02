#!/usr/bin/env python

# ^ shebang line is not necessary for Windows users, but it's great for Linux and Mac users.  It lets you run the script from the command line without having to type "python" before the script name.  You can just type "./example_pyxl.py" instead of "python example_pyxl.py"

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def create_cities_worksheet():
    # Create the workbook object
    wb = Workbook()
    # set the worksheet
    ws = wb.active
    ws.title = "Most Populous Cities"

    headers = ["Rank", "City", "Country", "Population", "Year", "% of Tokyo", "Population in Millions"]
    # underscore is a good way of showing the thousands separator, you can also just write out 100000 instead of 100_000
    data = [
        (1, "Tokyo", "Japan", 37_400_068, 2020),
        (2, "Delhi", "India", 32_941_000, 2020),
        (3, "Shanghai", "China", 27_796_000, 2020),
        (4, "São Paulo", "Brazil", 22_429_800, 2020),
        (5, "Mexico City", "Mexico", 22_085_140, 2020),
        (6, "Dhaka", "Bangladesh", 21_741_090, 2020),
        (7, "Cairo", "Egypt", 21_750_020, 2020),
        (8, "Beijing", "China", 20_896_820, 2020),
        (9, "Mumbai", "India", 20_411_274, 2020),
        (10, "Osaka", "Japan", 19_059_856, 2020),
    ]

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        # Set cell formatting
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = Alignment(horizontal="center")

    # Write data and formulas
    tokyo_population = data[0][3]  # Store Tokyo's population for percentage calculation
    for row, city_data in enumerate(data, 2):
        for col, value in enumerate(city_data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.alignment = Alignment(horizontal="center")
            
            if col == 4:  # Population column
                cell.number_format = "#,##0"
                
        # Add percentage formula
        percentage_cell = ws.cell(row=row, column=6)
        percentage_cell.value = f"=D{row}/D2*100"
        percentage_cell.number_format = "0.00%"
        
        # Add millions formula
        millions_cell = ws.cell(row=row, column=7)
        millions_cell.value = f"=D{row}/1000000"
        millions_cell.number_format = "#,##0.00"

    # Add totals row
    total_row = len(data) + 2
    ws.cell(row=total_row, column=1, value="Total").font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=f"=SUM(D2:D{total_row-1})").number_format = "#,##0"
    ws.cell(row=total_row, column=7, value=f"=SUM(G2:G{total_row-1})").number_format = "#,##0.00"

    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    wb.save("most_populous_cities.xlsx")

if __name__ == "__main__":
    create_cities_worksheet()
